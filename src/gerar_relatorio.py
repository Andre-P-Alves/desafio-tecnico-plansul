import logging
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_HEADER_BG = "1F4E79"
_HEADER_FG = "FFFFFF"
_SECAO_BG = "2E75B6"
_ALERTA_BG = "FFD700"
_APENAS_CSV_BG = "FFA500"
_ALTERNO_BG = "DCE6F1"

_COLUNAS_PT = {
    "num_guia": "ID Cobrança",
    "nome_beneficiario": "Nome (CSV)",
    "cpf_beneficiario": "CPF",
    "ans": "Cód. ANS (CSV)",
    "nome_operadora": "Convênio",
    "descricao_servico": "Procedimento (CSV)",
    "cod_tuss": "Cód. TUSS",
    "dt_realizacao": "Data Realização",
    "dt_lancamento": "Data Lançamento",
    "vl_servico": "Valor Bruto R$",
    "vl_glosa": "Glosa R$",
    "vl_liquido": "Valor Líquido R$",
    "paciente": "Nome (XLSX)",
    "registro_ans": "Cód. ANS (XLSX)",
    "procedimento": "Procedimento (XLSX)",
    "data_atendimento": "Data Atendimento",
    "valor": "Valor XLSX R$",
    "pdf_renomeado": "PDF Renomeado",
    "divergencias": "Divergências",
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_style(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(_HEADER_BG)
        cell.font = Font(bold=True, color=_HEADER_FG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def _ajustar_largura(ws):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 55)


def _escrever_df(ws, df: pd.DataFrame, header_row: int = 1, zebra: bool = True):
    headers = [_COLUNAS_PT.get(c, c) for c in df.columns]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _header_style(ws, header_row, len(headers))

    fill_alerta = _fill(_ALERTA_BG)
    fill_csv = _fill(_APENAS_CSV_BG)
    fill_alt = _fill(_ALTERNO_BG)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        div = str(row.get("divergencias", ""))
        if "apenas_csv" in div:
            row_fill = fill_csv
        elif div.strip():
            row_fill = fill_alerta
        elif zebra and r_idx % 2 == 0:
            row_fill = fill_alt
        else:
            row_fill = None

        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime()
                cell.number_format = "DD/MM/YYYY"
            elif isinstance(val, float) and pd.isna(val):
                cell.value = None
            else:
                cell.value = val
            if row_fill:
                cell.fill = row_fill

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _aba_resumo(ws, df: pd.DataFrame, alertas: list[dict], res_laudos: list[dict]):
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 22

    titulo = ws.cell(row=1, column=1, value="Relatório de Faturamento Hospitalar")
    titulo.font = Font(bold=True, color=_HEADER_FG, size=14)
    titulo.fill = _fill(_HEADER_BG)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:B1")

    renomeados = sum(1 for r in res_laudos if r["status"] == "renomeado")
    pulados = sum(1 for r in res_laudos if r["status"] == "pulado")
    erros = sum(1 for r in res_laudos if r["status"] == "erro")

    secoes: list[tuple[str | None, str, Any]] = [
        (None, "Cobranças", ""),
        ("sec", "Total de registros consolidados", len(df)),
        ("sec", "Registros em ambas as fontes", int(df["paciente"].notna().sum())),
        ("sec", "Apenas no CSV (não lançados internamente)", int(df["paciente"].isna().sum())),
        ("sec", "Total com alguma divergência", len(alertas)),
        (None, "Valores (R$)", ""),
        ("sec", "Valor líquido total (CSV)", round(float(df["vl_liquido"].sum()), 2)),
        ("sec", "Total de glosas (CSV)", round(float(df["vl_glosa"].sum()), 2)),
        ("sec", "Valor bruto total (XLSX)", round(float(df["valor"].dropna().sum()), 2)),
        (None, "Laudos", ""),
        ("sec", "Total de PDFs processados", len(res_laudos)),
        ("sec", "Renomeados com sucesso", renomeados),
        ("sec", "Pulados (sem correspondência ou destino já existe)", pulados),
        ("sec", "Erros", erros),
    ]

    row = 3
    for tipo, label, valor in secoes:
        if tipo is None:
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = Font(bold=True, color=_HEADER_FG)
            cell.fill = _fill(_SECAO_BG)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 20
            ws.merge_cells(f"A{row}:B{row}")
        else:
            lbl = ws.cell(row=row, column=1, value=label)
            lbl.font = Font(bold=True)
            val_cell = ws.cell(row=row, column=2, value=valor)
            val_cell.alignment = Alignment(horizontal="right")
            if row % 2 == 0:
                lbl.fill = _fill(_ALTERNO_BG)
                val_cell.fill = _fill(_ALTERNO_BG)
        row += 1


def _mapear_pdfs_por_cob(res_laudos: list[dict]) -> dict[str, str]:
    mapa: dict[str, str] = {}
    for r in res_laudos:
        if r["status"] == "renomeado" and r.get("motivo"):
            m = re.search(r"cob=(\S+)", r["motivo"])
            if m:
                mapa.setdefault(m.group(1), r["arquivo_destino"])
    return mapa


def gerar(
    df: pd.DataFrame,
    alertas: list[dict],
    res_laudos: list[dict],
    mes_ref: str,
    caminho_saida: str,
) -> str:
    """
    Gera relatório Excel com 3 abas: Resumo, Detalhamento e Alertas.
    Linhas com divergência são destacadas em amarelo; apenas_csv em laranja.
    """
    Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)

    df_det = df.copy()
    df_det["pdf_renomeado"] = df_det["num_guia"].map(_mapear_pdfs_por_cob(res_laudos))

    wb = Workbook()

    ws_res = wb.active
    ws_res.title = "Resumo"
    _aba_resumo(ws_res, df_det, alertas, res_laudos)

    ws_det = wb.create_sheet("Detalhamento")
    _escrever_df(ws_det, df_det)
    _ajustar_largura(ws_det)

    df_alertas = df_det[df_det["divergencias"].astype(str).str.strip() != ""].copy()
    ws_alt = wb.create_sheet("Alertas")
    if not df_alertas.empty:
        _escrever_df(ws_alt, df_alertas, zebra=False)
        _ajustar_largura(ws_alt)
    else:
        ws_alt.cell(row=1, column=1, value="Nenhuma divergência encontrada.")

    wb.save(caminho_saida)
    logger.info(
        "Relatório salvo: %s (%d alertas, %d registros)",
        caminho_saida, len(alertas), len(df_det),
    )
    return caminho_saida

"""Caso queira rodar o arquivo individualmente, printa no terminal o resultado"""

if __name__ == "__main__":
    import logging
    import sys
    import pandas as pd

    sys.path.insert(0, str(Path(__file__).parent))
    from consolidar_cobrancas import consolidar
    from renomear_laudos import renomear

    logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")

    _BASE = Path(__file__).parent.parent
    _DATA = _BASE / "data"

    _df, _alertas = consolidar(
        str(_DATA / "cobrancas_convenio.csv"),
        str(_DATA / "cobrancas_internas.xlsx"),
    )
    _res_laudos = renomear(_df, str(_DATA / "laudos"))

    _mes_ref = ""
    if not _df.empty and _df["dt_realizacao"].notna().any():
        _mes_ref = pd.Timestamp(_df["dt_realizacao"].dropna().iloc[0]).strftime("%m/%Y")
    _mes_arquivo = _mes_ref.replace("/", "") or "000000"

    _saida = str(_BASE / f"relatorio_faturamento_{_mes_arquivo}.xlsx")
    gerar(_df, _alertas, _res_laudos, _mes_ref, _saida)
    print(f"\nRelatório salvo em: {_saida}")
